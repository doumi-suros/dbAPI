import xlrd
import xlwt
import openpyxl
import pandas as pd


def keyword_to_uniqueDB():
	keyword = input("please input a keyword:")
	#keyword = '佛教慈濟醫療財團法人大林慈濟醫院'
	workbook = xlrd.open_workbook('DBsourceData_220801.xls') #讀取源excel
	result = xlwt.Workbook(encoding="utf-8")  #生成excel
	wsheet = result.add_sheet('RESULT') #生成sheet
	
	y=0 #生成的excel的行計數
	BaseData = workbook.sheet_by_name("BaseData") #讀取源excel第一張sheet(BaseData)內容
	nrowsnum = BaseData.nrows  #讀取第一張sheet(BaseData)的行數
	for i in range(0,nrowsnum):
		company=BaseData.row(i) #讀取第一張sheet(BaseData)第i行的内容
		for n in range(0,len(company)):
			companyrow=str(company[n]) #把該行第n個單元格轉化成字串，目的是下一步進行關鍵字比對
			if companyrow.find(keyword)>0: #進行關鍵字比對，包含關鍵字返回1，否則返回0
				y=y+1 
				for j in range(len(company)):
					if j==0: #DB編號
						wsheet.write(y,j,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==1: #中文全名
						wsheet.write(y,j,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==2: #中文簡稱
						wsheet.write(y,j,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==3: #英文全名
						wsheet.write(y,j,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==4: #英文簡稱
						wsheet.write(y,j,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==6: #別名1
						wsheet.write(y,j-1,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==7: #別名2
						wsheet.write(y,j-1,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==8: #別名3
						wsheet.write(y,j-1,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==10:#股票簡稱
						wsheet.write(y,j-2,BaseData.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行

	result.save('Totalresultlist.xls') #保存新生成的Excel


if __name__ == '__main__':
	keyword_to_uniqueDB()

resultworkbook = xlrd.open_workbook('Totalresultlist.xls')
sheet = resultworkbook.sheets()[0]
DBcodeA = sheet.col_values(0)
DBcodeAll=list(set(DBcodeA))
DBcodeAll.remove('')
print(DBcodeAll)


####DBkey
def DBcode_as_key():
	# key = DBcode ***尚未完成
	key = DBcodeAll[0]
	workbook = xlrd.open_workbook('DBsourceData_220801.xls') #讀取源excel
	Resultbook = xlrd.open_workbook('Totalresultlist.xls')
	BaseData = workbook.sheet_by_name("BaseData") #讀取源excel第一張sheet(BaseData)內容
	result = xlwt.Workbook(encoding="utf-8")  #生成excel
	wsheet = result.add_sheet('RESULT') #生成sheet
	
	wsheet.write(0,0,"DB編號")
	wsheet.write(0,1,"日期") 
	wsheet.write(0,2,"PDFFILE") 
	wsheet.write(0,3,"CSVFILE") 
	wsheet.write(0,4,"報告檔案夾") 

	y=0 #生成的excel的行計數
	Panorays_Rept = workbook.sheet_by_name("Panorays_Rept.List") #讀取源excel第三張sheet(Panorays_Rept.List)內容
	Totalresultlist = Resultbook.sheet_by_name("RESULT")
	nrowsnum = Panorays_Rept.nrows  #讀取第三張sheet(Panorays_Rept.List)的行數
	for i in range(0,nrowsnum):
		company=Panorays_Rept.row(i) #讀取第三張sheet(Panorays_Rept.List)第i行的内容
		for n in range(0,len(company)):
			companyrow=str(company[n]) #把該行第n個單元格轉化成字串，目的是下一步進行關鍵字比對
			if companyrow.find(key)>0: #進行關鍵字比對，包含關鍵字返回1，否則返回0
				y=y+1 
				for j in range(0,len(company)):
					if j==0: #DB編號
						wsheet.write(y,0,Panorays_Rept.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行
					if j==1: #報告檔案夾
						wsheet.write(y,4,Panorays_Rept.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行	
					if j==2: #日期
						wsheet.write(y,1,Panorays_Rept.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行	
					if j==3: #PDFFILE
						wsheet.write(y,2,Panorays_Rept.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行	
					if j==22:#CSVFILE
						wsheet.write(y,3,Panorays_Rept.cell_value(i,j)) #該行包含關鍵字，則把它所有單元格依次寫入新生成的excel的第y行				
						

	result.save('preHTML.xls') #保存新生成的Excel


if __name__ == '__main__':
	DBcode_as_key()

########[回BaseData，用DB編號補中文全名] ***尚未完成
df = pd.read_excel("./preHTML.xls",'RESULT')
dfC = pd.read_excel("./DBsourceData_220801.xls",'BaseData')
DBcode = list(df['DB編號'])
print(DBcode)
CHINESENAME = dfC['中文全名'].where(dfC['DB編號'] == DBcode[0])

########[用preHTML，將PDFFILE&CSVFILE=>PDF_PATH&CSV_FILE]

Pathhead ="C:/SurosBI/Panorays_Reports_220801A"
PDFPATH = list(Pathhead + "/" + df['報告檔案夾']+ "/" + df['PDFFILE'])
CSVPATH = list(Pathhead + "/" + df['報告檔案夾']+ "/" + df['CSVFILE'])

df.insert(1, column="中文全名", value = CHINESENAME)
df.insert(6, column="PDF_PATH", value = PDFPATH)
df.insert(7, column="CSV_PATH", value = CSVPATH)

file_name = 'HTML.xls'
HTML = df.to_excel(file_name)

# def PATH():
# 	workbook = xlrd.open_workbook('preHTML.xls') #讀取源excel
# 	activesheet = workbook.sheet_by_name("RESULT")
# 	result = xlwt.Workbook(encoding="utf-8")  #生成excel
# 	wsheet = workbook.active
	
# 	wsheet.write(0,0,"DB編號")
# 	wsheet.write(0,1,"中文全名") 
# 	wsheet.write(0,2,"日期") 
# 	wsheet.write(0,3,"PDF_PATH") 
# 	wsheet.write(0,4,"CSV_PATH") 
	
# 	df = pd.read_excel("./preHTML.xls",'RESULT')
# 	df['Full PDF_PATH'] = df["報告檔案夾"] + "/" + df["PDFFILE"]


# 	result.save('HTML.xls') #保存新生成的Excel


# if __name__ == '__main__':
# 	PATH()

# ######合併data of 3rd =>print out the path 
# DBlength = len(DBcodeAll)

# df3 = pd.read_excel("./DBsourceData_220801.xlsx",'Panorays_Rept.List')

# PDFFileAll=list(range(DBlength)) #for迴圈(givenDB)的所有PDF結果放入的List
# CSVFileAll=list(range(DBlength)) #for迴圈(givenDB)的所有CSV結果放入的List
# i=0
# for i in range(1, DBlength):
# 	givenDB=DBcodeAll[i]
# 	PathFolder= df3['報告檔案夾'].where(df3['DB編號'] == givenDB )
# 	PathPDF= df3['pdfName'].where(df3['DB編號'] == givenDB )
# 	PathCSV= df3['csvName'].where(df3['DB編號'] == givenDB )
# 	New_PathFolder=PathFolder.dropna(axis=0,how='any')
# 	New_PathPDF=PathPDF.dropna(axis=0,how='any')
# 	New_PathCSV=PathCSV.dropna(axis=0,how='any')
# 	#Folder = New_PathFolder.tolist()[0]

# 	if len(New_PathFolder.tolist())==0:
# 		print("there is no folder for this company.")
# 	else:
# 		Folder = New_PathFolder.tolist()[0]

# 	PDFFile = New_PathPDF.tolist() 
# 	CSVFile = New_PathCSV.tolist()

# 	Pathhead ="C:/SurosBI/Panorays_Reports_220801A"
# 	PDFPath = [Pathhead +"/"+ Folder + "/" + str(i) for i in PDFFile]
# 	CSVPath = [Pathhead +"/"+ Folder + "/" + str(i) for i in CSVFile]

# 	PDFFileAll[i]=PDFFileAll.extend(PDFPath)
# 	CSVFileAll[i]=CSVFileAll.extend(CSVPath)
# 	i=i+1

# ClearedPDFFileAll=list(filter(None, PDFFileAll))
# ClearedCSVFileAll=list(filter(None, CSVFileAll))

# print(ClearedPDFFileAll)
# print(ClearedCSVFileAll)

# # with open("C:/SurosBI/Panorays_Reports_220801A/台視_ttv.com.tw/ttv.com.tw_findings_Jun_17_2022.csv",'r',encoding='utf-8') as f:
# #  	print(f.read())

# # with pdfplumber.open("C:/SurosBI/Panorays_Reports_220801A/台視_ttv.com.tw/ttv.com.tw_summary_Jun_17_2022.pdf") as pdf:
# # 	print(pdf)	

# ################