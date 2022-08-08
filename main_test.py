import sub_excel
import sub_msc


searchKey = '統一'
#searchCol = 'B'
searchCol = 2



sub_excel.sSearch (searchKey, searchCol)



"""
from openpyxl import load_workbook, Workbook


exlWB = load_workbook('DBsourceData_220804.xlsx')    #get excel file / workbook
exlST = exlWB['BaseData']    #get file work sheet

newWB = Workbook()
newST = newWB.active

newHeaders = ['公司全名', '日期', 'PDF', 'CSV']
newST.append(newHeader)




flag = 0
for cell in exlST:
    if cell.value == searchKey:
        flag = cell.row
        break
"""



