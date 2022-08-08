from openpyxl import load_workbook, Workbook

exlWB = load_workbook('DBsourceData_220804.xlsx')    #get excel file / workbook
exlST = exlWB['BaseData']    #get file work sheet


#by search key to search single column 
def sSearch (searchKey, searchCol):
    for searchR in range (1, exlST.max_row + 1):
        for searchC in range (1, exlST.max_column + 1):
            a = exlST.cell(row = searchR, column = searchC)
            if a == searchKey:
                print (a, 'ROW '+str(searchR), 'COL '+str(searchC))


